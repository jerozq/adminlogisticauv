# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path

path = Path('2. REQUERIMIENTO EXCEL 629PE.xlsx')
wb = openpyxl.load_workbook(path, data_only=True)

# Hojas clave a analizar en detalle
target_sheets = []
for name in wb.sheetnames:
    n = name.upper()
    if any(k in n for k in ['FORMATO', 'MATERIAL', 'DETALLE', 'ESPECIF', 'ALOJAM', 'TRANSPORT']):
        target_sheets.append(name)

print('Hojas objetivo:', target_sheets)

for sname in target_sheets:
    ws = wb[sname]
    print(f'\n{"="*60}')
    print(f'HOJA: {sname}  (filas={ws.max_row}, cols={ws.max_column})')
    print(f'{"="*60}')
    for r in range(1, min(ws.max_row+1, 120)):
        row_vals = []
        for c in range(1, min(ws.max_column+1, 20)):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                row_vals.append(f'  C{c}={repr(str(v).strip()[:60])}')
        if row_vals:
            print(f'R{r:03d}:{"".join(row_vals)}')
